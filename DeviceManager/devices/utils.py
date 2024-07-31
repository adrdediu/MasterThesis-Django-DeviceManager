# devices/utils.py

import io
import json
from django.template.loader import get_template
from django.db.models import Count, Q
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from .models import Room, Device, DeviceScan

def generate_inventory_pdf_report(inventory):
    # Get the template
    template = get_template('devices/inventory_report_pdf.html')
    
    # Load the JSON data
    with open(inventory.inventory_data_file.path, 'r') as f:
        data = json.load(f)
    
    # Prepare the context
    context = {
        'inventory': data,
        'rooms_data': [],
        'total_devices': data['total_devices'],
        'total_scanned': data['total_scanned']
    }
    
    # Group devices by room
    rooms = {}
    for device in data['devices']:
        room_id = device['room']['id']
        if room_id not in rooms:
            rooms[room_id] = {
                'room': device['room'],
                'devices': [],
                'scanned_devices': []
            }
        rooms[room_id]['devices'].append(device)
    
    # Add scanned devices
    for scan in data['device_scans']:
        for room in rooms.values():
            for device in room['devices']:
                if device['id'] == scan['device_id']:
                    room['scanned_devices'].append(device)
                    break
    
    context['rooms_data'] = list(rooms.values())

    # Render the template
    html = template.render(context)

    # Create a file-like buffer to receive PDF data.
    buffer = io.BytesIO()

    # Generate the PDF
    pisa_status = pisa.CreatePDF(html, dest=buffer)

    # If error creating PDF, return None
    if pisa_status.err:
        return None
    
    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)
    return buffer

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import json
import os
from django.conf import settings
import logging

logger = logging.getLogger(__name__)

def generate_inventory_excel_report(inventory):
    try:
        json_file_path = os.path.join(settings.MEDIA_ROOT, inventory.inventory_data_file.name)
        with open(json_file_path, 'r') as f:
            data = json.load(f)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Inventory Report {data['inventorization_id']}"

        # General Information
        ws['A1'] = "General Information"
        ws['A1'].font = Font(bold=True, size=14)
        
        info_fields = ["inventorization_id", "creator", "inventory", "building", "start_date", "end_date", "status", "total_devices", "total_scanned"]
        for i, field in enumerate(info_fields, start=2):
            ws[f'A{i}'] = field.replace('_', ' ').title()
            ws[f'B{i}'] = str(data.get(field, ""))

        # Devices
        ws['A12'] = "Devices"
        ws['A12'].font = Font(bold=True, size=14)

        headers = ["ID", "Name", "Serial Number", "Category", "Subcategory", "Owner", "Building", "Floor", "Room", "Status"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=13, column=col, value=header).font = Font(bold=True)

        row = 14
        for device in data['devices']:
            for col, field in enumerate(["id", "name", "serial_number", "category.name", "subcategory.name", "owner", "building.name", "floor.name", "room.name"], start=1):
                value = device
                for key in field.split('.'):
                    value = value.get(key, {})
                ws.cell(row=row, column=col, value=str(value))
            ws.cell(row=row, column=10, value="Scanned" if any(scan['device_id'] == device['id'] for scan in data['device_scans']) else "Not Scanned")
            row += 1

        # Changes
        ws.cell(row=row + 2, column=1, value="Changes").font = Font(bold=True, size=14)

        headers = ["Device ID", "Change Type", "Timestamp", "User"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=row + 3, column=col, value=header).font = Font(bold=True)

        for i, change in enumerate(data['changes'], start=row + 4):
            for col, field in enumerate(["device_id", "change_type", "timestamp", "user"], start=1):
                ws.cell(row=i, column=col, value=str(change.get(field, "")))

        # Adjust column widths
        for col in range(1, 11):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Save the workbook
        excel_file = f'inventory_report_{data["inventorization_id"]}.xlsx'
        excel_file_path = os.path.join(settings.MEDIA_ROOT, 'inventory_reports', excel_file)
        os.makedirs(os.path.dirname(excel_file_path), exist_ok=True)
        wb.save(excel_file_path)
        logger.info(f"Excel report generated successfully: {excel_file_path}")
        return excel_file_path
    except Exception as e:
        logger.error(f"Error generating Excel report: {str(e)}")
        raise

